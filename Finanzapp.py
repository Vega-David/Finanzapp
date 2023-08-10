from datetime import datetime
import calendar
import openpyxl
import os

class RegistroFinanciero:
    def __init__(self, archivo_excel):
        self.archivo_excel = archivo_excel
        self.registro = []
        self.cargar_registro()

    def cargar_registro(self):
        try:
            # Intenta cargar el archivo Excel existente
            workbook = openpyxl.load_workbook(self.archivo_excel)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Agrega cada transacción a la lista de registros
                self.registro.append({
                    'Fecha': row[0],
                    'Descripción': row[1],
                    'Ingresos': row[2],
                    'Gastos': row[3]
                })
            workbook.close()
        except FileNotFoundError:
            print("Archivo no encontrado. Se creará uno nuevo.")

    def guardar_registro(self):
        # Crea un nuevo archivo Excel y guarda los registros en él
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "Descripción", "Ingresos", "Gastos"])
        for transaccion in self.registro:
            sheet.append([transaccion['Fecha'], transaccion['Descripción'], transaccion['Ingresos'], transaccion['Gastos']])
        workbook.save(self.archivo_excel)

    def agregar_transaccion(self, fecha, descripcion, ingresos, gastos):
        # Agrega una nueva transacción a la lista de registros
        self.registro.append({
            'Fecha': fecha,
            'Descripción': descripcion,
            'Ingresos': ingresos,
            'Gastos': gastos
        })

    def separar_gastos_por_mes(self):
        gastos_por_mes = {}
        for transaccion in self.registro:
            # Separa las transacciones por mes y año
            fecha = datetime.strptime(transaccion['Fecha'], '%d-%m-%Y')
            mes = fecha.month
            anio = fecha.year
            if (mes, anio) in gastos_por_mes:
                gastos_por_mes[(mes, anio)].append(transaccion)
            else:
                gastos_por_mes[(mes, anio)] = [transaccion]
        return gastos_por_mes

    def mostrar_transacciones_mes(self):
        gastos_por_mes = self.separar_gastos_por_mes()

        os.system('cls' if os.name == 'nt' else 'clear')
        print("\nMESES DISPONIBLES:")
        for idx, (mes, anio) in enumerate(gastos_por_mes.keys(), start=1):
            # Muestra los meses disponibles para visualizar
            print(f"{idx}. {calendar.month_name[mes]} {anio}")

        opcion = input("Seleccione el número de mes a visualizar (1, 2, ...): ")
        try:
            opcion = int(opcion)
            if opcion < 1 or opcion > len(gastos_por_mes):
                print("Selección inválida.")
                return
        except ValueError:
            print("Entrada inválida. Debe ingresar un número.")
            return

        mes_seleccionado = list(gastos_por_mes.keys())[opcion - 1]
        transacciones = gastos_por_mes[mes_seleccionado]

        os.system('cls' if os.name == 'nt' else 'clear')
        print(f"\nTRANSACCIONES DE {calendar.month_name[mes_seleccionado[0]]} {mes_seleccionado[1]}:")
        for idx, transaccion in enumerate(transacciones, start=1):
            # Muestra las transacciones del mes seleccionado
            print(f"{idx}. Fecha: {transaccion['Fecha']}, Descripción: {transaccion['Descripción']}, Ingresos: {transaccion['Ingresos']}, Gastos: {transaccion['Gastos']}")

        opcion_transaccion = input("\nSeleccione el número de transacción a modificar (1, 2, ...) o presione Enter para volver al menú: ")
        if opcion_transaccion:
            try:
                opcion_transaccion = int(opcion_transaccion)
                if opcion_transaccion < 1 or opcion_transaccion > len(transacciones):
                    print("Selección inválida.")
                    return
            except ValueError:
                print("Entrada inválida. Debe ingresar un número.")
                return

            transaccion_seleccionada = transacciones[opcion_transaccion - 1]

            os.system('cls' if os.name == 'nt' else 'clear')
            print(f"\nMODIFICANDO TRANSACCIÓN:")
            print(f"Fecha: {transaccion_seleccionada['Fecha']}")
            print(f"Descripción: {transaccion_seleccionada['Descripción']}")
            print(f"Ingresos: {transaccion_seleccionada['Ingresos']}")
            print(f"Gastos: {transaccion_seleccionada['Gastos']}")

            print("\nIngrese los nuevos datos:")
            descripcion = input("Ingrese una nueva descripción: ")
            ingresos = float(input("Ingrese los nuevos ingresos: "))
            gastos = float(input("Ingrese los nuevos gastos: "))

            registro_actualizado = {
                'Fecha': transaccion_seleccionada['Fecha'],
                'Descripción': descripcion,
                'Ingresos': ingresos,
                'Gastos': gastos
            }

            transacciones[opcion_transaccion - 1] = registro_actualizado
            self.guardar_registro()

            print("\nTransacción modificada exitosamente.")

    def mostrar_gastos_por_mes(self):
        gastos_por_mes = self.separar_gastos_por_mes()
        for key, transacciones in gastos_por_mes.items():
            total_ingresos = sum(transaccion['Ingresos'] for transaccion in transacciones)
            total_gastos = sum(transaccion['Gastos'] for transaccion in transacciones)
            saldo_mensual = total_ingresos - total_gastos
            print(f"\nMes: {calendar.month_name[key[0]]} {key[1]}")
            print("Total Ingresos:", total_ingresos)
            print("Total Gastos:", total_gastos)
            print("Saldo Mensual:", saldo_mensual)
            if saldo_mensual < 0:
                print("Deuda para el siguiente mes:", abs(saldo_mensual))

def main():
    archivo_excel = "registro_financiero.xlsx"
    registro = RegistroFinanciero(archivo_excel)

    nombre_usuario = ""
    try:
        with open("nombre_usuario.txt", "r") as f:
            nombre_usuario = f.read().strip()
    except FileNotFoundError:
        nombre_usuario = input("Hola, ¿cuál es tu nombre? ")
        with open("nombre_usuario.txt", "w") as f:
            f.write(nombre_usuario)

    os.system('cls' if os.name == 'nt' else 'clear')
    print(f"Hola {nombre_usuario}!")
    input("Presiona Enter para continuar...")

    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        print("\nMENU:")
        print("1. Agregar Transacción")
        print("2. Mostrar Gastos por Mes y Modificar Transacción")
        print("3. Guardar y Salir")

        opcion = input("Selecciona una opción: ")

        if opcion == '1':
            os.system('cls' if os.name == 'nt' else 'clear')
            descripcion = input("Ingrese una descripción: ")
            ingresos = float(input("Ingrese los ingresos: "))
            gastos = float(input("Ingrese los gastos: "))

            hoy = input("¿La transacción es del día de hoy? (s/n): ")
            if hoy.lower() == 's':
                fecha = datetime.now().strftime('%d-%m-%Y')
            else:
                fecha = input("Ingrese la fecha (dd-mm-aaaa): ")

            registro.agregar_transaccion(fecha, descripcion, ingresos, gastos)

        elif opcion == '2':
            registro.mostrar_transacciones_mes()

        elif opcion == '3':
            registro.guardar_registro()
            break

        else:
            print("Opción no válida. Por favor, elige una opción del menú.")

if __name__ == "__main__":
    main()
    